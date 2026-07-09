#!/usr/bin/env python3
"""Convert the USGS earthquake CSV into a realistic .xlsx workbook
(typed columns, a header style, an extra summary sheet).

Usage: make_xlsx.py <earthquakes.csv> <out.xlsx>
"""
import csv
import sys

from openpyxl import Workbook
from openpyxl.styles import Font

NUMERIC = {"latitude", "longitude", "depth", "mag", "nst", "gap", "dmin", "rms",
           "horizontalError", "depthError", "magError", "magNst"}


def main(src: str, dst: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Earthquakes"
    with open(src, newline="", encoding="utf-8") as f:
        reader = csv.reader(f)
        header = next(reader)
        ws.append(header)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        idx_numeric = {i for i, name in enumerate(header) if name in NUMERIC}
        mags = []
        for row in reader:
            typed = []
            for i, value in enumerate(row):
                if i in idx_numeric and value:
                    try:
                        value = float(value)
                    except ValueError:
                        pass
                typed.append(value)
            ws.append(typed)
            try:
                mags.append(float(row[4]))
            except (ValueError, IndexError):
                pass

    summary = wb.create_sheet("Summary")
    summary.append(["Metric", "Value"])
    summary.append(["Events", len(mags)])
    summary.append(["Max magnitude", max(mags)])
    summary.append(["Min magnitude", min(mags)])
    summary.append(["Mean magnitude", sum(mags) / len(mags)])
    for cell in summary[1]:
        cell.font = Font(bold=True)

    wb.save(dst)


if __name__ == "__main__":
    main(sys.argv[1], sys.argv[2])
